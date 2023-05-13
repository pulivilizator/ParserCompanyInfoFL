import openpyxl
import configparser
from exeptions import GetFileExeption, WriteFileExeption

def _create_file(config):
    road = config.get("program", "write_in_file")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(['Тип', 'regNumber', 'Регистрационный номер поставщика', 'Полное название', 'ИНН',
                      'Телефон в контрактах, число раз',
                      'Телефон в контрактах', 'Телефон', 'Комментарий', 'КПП', 'Доход 2021', 'Доход 2022',
                      'Налог 2021', 'Налог 2020', 'ОГРН', 'largestTaxpayerKpp',
                      'МСП', 'Состояние организации', 'okved', 'supplierOkved', 'Адрес', 'Почтовый адрес',
                      'Дата регистрации (юрлица?)',
                      'Дата регистрации поставщика', 'isSMP', 'Клиент', 'Поставщик',
                      'Эксклюзивный поставщик', 'Уровень подчинения', 'Описание уровня подчинения',
                      'Недобросовестный ФЗ44 поставщик',
                      'Недобросовестный ФЗ233 поставщик', 'isSono', 'Самозанятый', 'operatorEdo', 'abonentId',
                      'Директор', 'email', None, 'Факс',
                      'Сайт', 'Запрещено подавать заявки', 'Причина запрета подачи заявок', 'Запрещено делать закупки',
                      'Причина запрета закупок',
                      'id', None, None, None, None])
    worksheet.auto_filter.ref = 'A1:AT1'
    workbook.save(road)
    print(f'Создан файл по пути {road}')


def _get_rows(config):
    try:
        road = config.get("program", "read_file")
        rows = []
        workbook = openpyxl.load_workbook(road)

        worksheet = workbook.active

        for row in worksheet.iter_rows():
            rows.append([i.value for i in row])
    except FileNotFoundError:
        print(f'GetFileExeption: {GetFileExeption.__doc__}')
        raise GetFileExeption
    return rows


def _writer(row):
    try:
        config = configparser.ConfigParser()
        config.read('config.ini', encoding='utf-8')
        road = config.get("program", "write_in_file")
        workbook = openpyxl.load_workbook(road)

        worksheet = workbook.active

        worksheet.append(row)

        # Сохраняем файл
        workbook.save(road)
    except FileNotFoundError:
        print(f'WriteFileExeption: {WriteFileExeption.__doc__}')
        raise WriteFileExeption
